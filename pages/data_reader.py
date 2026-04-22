import openpyxl

def read_excel(filename, sheet_name):
    wb = openpyxl.load_workbook(filename)
    ws = wb[sheet_name]
    
    headers = [cell.value for cell in ws[1]]
    data = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            # Convert None to empty string
            clean_row = tuple("" if cell is None else cell for cell in row)
            data.append(dict(zip(headers, clean_row)))
    
    return data